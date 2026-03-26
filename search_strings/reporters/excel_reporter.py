"""Excel report generator."""

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pandas import ExcelWriter
from ..config import (
    EXCEL_HEADER_COLOR,
    EXCEL_FOUND_COLOR,
    EXCEL_MISSING_COLOR,
    EXCEL_MONOSPACE_FONT,
    EXCEL_MONOSPACE_SIZE,
)
from .base import Reporter


class ExcelReporter(Reporter):
    """Generate Excel reports with optimized formatting for long paths."""

    def generate(self, results: list, pattern_line_hits: dict, output_file: str, pandas=None, openpyxl=None, **kwargs):
        """
        Generate an Excel report with three sheets:
        - Summary: Pattern status overview
        - Results: Files per pattern
        - Lines: Detailed line-by-line matches
        """
        if pandas is None or openpyxl is None:
            raise ValueError("pandas and openpyxl are required for Excel reporting")

        # Prepare dataframes
        df_sum = pandas.DataFrame(
            [
                {"pattern": r["pattern"], "status": r["status"], "count": len(r["paths"])}
                for r in results
            ]
        )

        df_res = pandas.DataFrame(
            [
                {"pattern": r["pattern"], "status": r["status"], "path": p}
                for r in results
                for p in r["paths"]
            ]
        )

        df_lines = pandas.DataFrame([
            entry
            for plist in pattern_line_hits.values()
            for entry in plist
        ])

        with ExcelWriter(output_file, engine="openpyxl") as writer:
            # Summary sheet
            df_sum.to_excel(writer, index=False, sheet_name="Summary")
            self._format_summary_sheet(writer.sheets["Summary"], df_sum)

            # Results sheet
            df_res.to_excel(writer, index=False, sheet_name="Results")
            self._format_results_sheet(writer.sheets["Results"], df_res)

            # Lines sheet
            df_lines.to_excel(writer, index=False, sheet_name="Lines")
            self._format_lines_sheet(writer.sheets["Lines"], df_lines)

    def _format_summary_sheet(self, ws, df):
        """Format the Summary sheet."""
        header_fill = PatternFill(start_color=EXCEL_HEADER_COLOR, fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col in range(1, len(df.columns) + 1):
            c = ws.cell(row=1, column=col)
            c.fill = header_fill
            c.font = header_font
            ws.column_dimensions[get_column_letter(col)].width = 20

        # Row coloring
        for row in range(2, len(df) + 2):
            c = ws.cell(row=row, column=2)
            if c.value == "FOUND":
                c.fill = PatternFill(start_color=EXCEL_FOUND_COLOR, fill_type="solid")
            else:
                c.fill = PatternFill(start_color=EXCEL_MISSING_COLOR, fill_type="solid")

    def _format_results_sheet(self, ws, df):
        """Format the Results sheet with monospace font for paths."""
        header_fill = PatternFill(start_color=EXCEL_HEADER_COLOR, fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        monospace_font = Font(name=EXCEL_MONOSPACE_FONT, size=EXCEL_MONOSPACE_SIZE)

        for col in range(1, 4):
            c = ws.cell(row=1, column=col)
            c.fill = header_fill
            c.font = header_font
            ws.column_dimensions[get_column_letter(col)].width = 50

        for row in range(2, len(df) + 2):
            cell = ws.cell(row=row, column=3)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font = monospace_font
            ws.row_dimensions[row].height = 45

    def _format_lines_sheet(self, ws, df):
        """Format the Lines sheet with monospace fonts and optimal sizing."""
        header_fill = PatternFill(start_color=EXCEL_HEADER_COLOR, fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        monospace_font = Font(name=EXCEL_MONOSPACE_FONT, size=EXCEL_MONOSPACE_SIZE)

        # Header styling
        for col in range(1, len(df.columns) + 1):
            c = ws.cell(row=1, column=col)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(wrap_text=True, vertical="top")

        # Column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 70
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 100

        # Row formatting
        for row in range(2, len(df) + 2):
            # Pattern column (A)
            ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
            # File path column (B) - monospace
            ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=row, column=2).font = monospace_font
            # Line number column (C) - centered
            ws.cell(row=row, column=3).alignment = Alignment(horizontal="center", vertical="top")
            # Content column (D) - monospace
            ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=row, column=4).font = monospace_font
            ws.row_dimensions[row].height = 50
