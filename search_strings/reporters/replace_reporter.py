"""Replace report generators for HTML, Excel, CSV, JSON output."""

import csv
import json
from datetime import datetime
from collections import Counter

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pandas import ExcelWriter

from ..config import (
    EXCEL_HEADER_COLOR,
    EXCEL_MONOSPACE_FONT,
    EXCEL_MONOSPACE_SIZE,
    REPLACE_HTML_TEMPLATE,
)


class ReplaceHtmlReporter:
    """Generate HTML report for replace operations."""

    def generate(self, changes: list, skipped: list, files_scanned: int,
                 dry_run: bool, output_file: str):
        html = REPLACE_HTML_TEMPLATE.replace(
            "{{DATE}}", datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        )
        mode = "Dry Run Preview" if dry_run else "Replacements Applied"
        html = html.replace("{{MODE}}", mode)

        # Summary stats
        files_affected = len({c["file"] for c in changes})
        pair_counts = Counter()
        for c in changes:
            pair_counts[(c["old"], c["new"])] += 1

        summary = f"Files scanned: {files_scanned} | "
        summary += f"Files affected: {files_affected} | "
        summary += f"Total replacements: {len(changes)}"
        if skipped:
            summary += f" | Skipped: {len(skipped)}"
        html = html.replace("{{SUMMARY}}", summary)

        # Pair summary
        pair_blocks = []
        for (old, new), count in pair_counts.items():
            new_display = new if new else "(deletion)"
            pair_blocks.append(
                f'<div class="pair-item">'
                f'<span class="old-str">{_escape(old)}</span> -&gt; '
                f'<span class="new-str">{_escape(new_display)}</span> '
                f'<span class="count-badge">{count}</span></div>'
            )
        html = html.replace("{{PAIRS}}", "\n".join(pair_blocks))

        # Changes grouped by file
        files_map = {}
        for c in changes:
            files_map.setdefault(c["file"], []).append(c)

        blocks = []
        for filepath, file_changes in files_map.items():
            part = f"""
        <details open>
            <summary>
                <span class="file-name">{_escape(filepath)}</span>
                <span class="count-badge">{len(file_changes)}</span>
            </summary>
            <div class="change-list">
        """
            for c in file_changes:
                part += f"""
                <div class="change-item">
                    <div class="line-num">Line {c['line']}</div>
                    <div class="diff-line removed">--- {_escape(c['content_before'])}</div>
                    <div class="diff-line added">+++ {_escape(c['content_after'])}</div>
                </div>
        """
            part += "</div></details>"
            blocks.append(part)

        html = html.replace("{{CONTENT}}", "\n".join(blocks))

        with open(output_file, "w", encoding="utf-8") as f:
            f.write(html)


class ReplaceExcelReporter:
    """Generate Excel report for replace operations."""

    def generate(self, changes: list, skipped: list, files_scanned: int,
                 dry_run: bool, output_file: str, pandas=None, openpyxl=None):
        if pandas is None or openpyxl is None:
            raise ValueError("pandas and openpyxl are required for Excel reporting")

        # Summary sheet data
        files_affected = len({c["file"] for c in changes})
        pair_counts = Counter()
        for c in changes:
            pair_counts[(c["old"], c["new"])] += 1

        summary_rows = [
            {"metric": "Files scanned", "value": files_scanned},
            {"metric": "Files affected", "value": files_affected},
            {"metric": "Total replacements", "value": len(changes)},
            {"metric": "Files skipped", "value": len(skipped)},
            {"metric": "Mode", "value": "Dry Run" if dry_run else "Applied"},
        ]
        df_summary = pandas.DataFrame(summary_rows)

        # Pairs sheet
        pair_rows = [
            {"old_string": old, "new_string": new, "matches": count}
            for (old, new), count in pair_counts.items()
        ]
        df_pairs = pandas.DataFrame(pair_rows) if pair_rows else pandas.DataFrame(
            columns=["old_string", "new_string", "matches"]
        )

        # Changes sheet
        df_changes = pandas.DataFrame([
            {
                "file": c["file"],
                "line": c["line"],
                "old_string": c["old"],
                "new_string": c["new"],
                "before": c["content_before"],
                "after": c["content_after"],
                "applied": c["applied"],
            }
            for c in changes
        ]) if changes else pandas.DataFrame(
            columns=["file", "line", "old_string", "new_string",
                      "before", "after", "applied"]
        )

        with ExcelWriter(output_file, engine="openpyxl") as writer:
            df_summary.to_excel(writer, index=False, sheet_name="Summary")
            self._format_header(writer.sheets["Summary"], df_summary)

            df_pairs.to_excel(writer, index=False, sheet_name="Pairs")
            self._format_header(writer.sheets["Pairs"], df_pairs)

            df_changes.to_excel(writer, index=False, sheet_name="Changes")
            self._format_changes(writer.sheets["Changes"], df_changes)

    def _format_header(self, ws, df):
        header_fill = PatternFill(start_color=EXCEL_HEADER_COLOR, fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col in range(1, len(df.columns) + 1):
            c = ws.cell(row=1, column=col)
            c.fill = header_fill
            c.font = header_font
            ws.column_dimensions[get_column_letter(col)].width = 25

    def _format_changes(self, ws, df):
        header_fill = PatternFill(start_color=EXCEL_HEADER_COLOR, fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        monospace = Font(name=EXCEL_MONOSPACE_FONT, size=EXCEL_MONOSPACE_SIZE)

        for col in range(1, len(df.columns) + 1):
            c = ws.cell(row=1, column=col)
            c.fill = header_fill
            c.font = header_font

        widths = {"A": 60, "B": 8, "C": 25, "D": 25, "E": 80, "F": 80, "G": 10}
        for letter, w in widths.items():
            ws.column_dimensions[letter].width = w

        for row in range(2, len(df) + 2):
            for col in [1, 5, 6]:
                cell = ws.cell(row=row, column=col)
                cell.font = monospace
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[row].height = 45


class ReplaceJsonReporter:
    """Generate JSON report for replace operations."""

    def generate(self, changes: list, skipped: list, files_scanned: int,
                 dry_run: bool, output_file: str):
        report = {
            "mode": "dry_run" if dry_run else "applied",
            "files_scanned": files_scanned,
            "files_affected": len({c["file"] for c in changes}),
            "total_replacements": len(changes),
            "skipped": skipped,
            "changes": changes,
        }
        with open(output_file, "w", encoding="utf-8") as f:
            json.dump(report, f, indent=4)


class ReplaceCsvReporter:
    """Generate CSV report for replace operations."""

    def generate(self, changes: list, skipped: list, files_scanned: int,
                 dry_run: bool, output_file: str):
        with open(output_file, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["file", "line", "old_string", "new_string",
                         "content_before", "content_after", "applied"])
            for c in changes:
                w.writerow([
                    c["file"], c["line"], c["old"], c["new"],
                    c["content_before"], c["content_after"], c["applied"],
                ])


def _escape(text):
    """Escape HTML special characters."""
    return (text
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;"))
